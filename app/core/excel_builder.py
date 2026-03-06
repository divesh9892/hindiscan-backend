import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from app.core.logger import log
from app.core.font_converter import unicode_to_krutidev

class ExcelBuilder:
    def __init__(self, json_path, output_path="output_report.xlsx", use_legacy_font=False, legacy_font_name="Kruti Dev 010"):
        self.json_path = json_path
        self.output_path = output_path
        self.use_legacy_font = use_legacy_font
        self.legacy_font_name = legacy_font_name
        self.wb = Workbook()
        # Initialize the first sheet (we will name it dynamically in build())
        self.ws = self.wb.active 
        self.current_row = 1
        
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.header_fill = PatternFill(start_color="EAEAEA", end_color="EAEAEA", fill_type="solid")
        self.center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        self.left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    def _get_font(self, size, is_bold):
        if self.use_legacy_font:
            return Font(name=self.legacy_font_name, size=size + 2, bold=is_bold) 
        return Font(name="Nirmala UI", size=size, bold=is_bold)

    def _process_text(self, text):
        if self.use_legacy_font and isinstance(text, str):
            return unicode_to_krutidev(text)
        return text

    def load_data(self):
        if not os.path.exists(self.json_path):
            raise FileNotFoundError(f"❌ Could not find {self.json_path}. Please create it first.")
        with open(self.json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
            
        # 🚀 Detects Multi-Page Array vs Single Document (for backward compatibility)
        if "pages" in data:
            return [page.get("document", {}) if "document" in page else page for page in data["pages"]]
        elif "document" in data:
            return [data.get("document", {})]
        return []

    def get_max_columns(self, document):
        max_cols = 1
        for table in document.get("tables", []):
            if "headers" in table:
                max_cols = max(max_cols, len(table["headers"]))
        return max_cols

    def write_merged_text(self, text, max_cols, is_bold, font_size, alignment):
        if not text:
            return
        
        processed_text = self._process_text(text)
        cell = self.ws.cell(row=self.current_row, column=1, value=processed_text)
        self.ws.merge_cells(start_row=self.current_row, start_column=1, end_row=self.current_row, end_column=max_cols)
        
        cell.font = self._get_font(size=font_size, is_bold=is_bold)
        cell.alignment = alignment
        
        chars_per_line = max(max_cols * 15, 30) 
        estimated_lines = str(processed_text).count('\n') + (len(str(processed_text)) // chars_per_line) + 1
        self.ws.row_dimensions[self.current_row].height = estimated_lines * (font_size * 1.5)
        self.current_row += 1

    def _autofit_columns(self):
        merged_cells_map = set()
        for merged_range in self.ws.merged_cells.ranges:
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    merged_cells_map.add((row, col))

        max_col_idx = self.ws.max_column
        max_row_idx = self.ws.max_row
        
        for col_idx in range(1, max_col_idx + 1):
            column_letter = get_column_letter(col_idx)
            max_length = 0
            
            for row_idx in range(1, max_row_idx + 1):
                if (row_idx, col_idx) in merged_cells_map:
                    continue
                
                cell = self.ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    lines = str(cell.value).split('\n')
                    longest_line = max([len(line) for line in lines])
                    if longest_line > max_length:
                        max_length = longest_line
            
            adjusted_width = max(12, min(max_length + 4, 45))
            self.ws.column_dimensions[column_letter].width = adjusted_width

    def build(self):
        log.info("🚀 Booting Smart Excel Builder...")
        pages_data = self.load_data()
        
        if not pages_data:
            log.error("❌ Invalid JSON format or empty document.")
            return

        # 🚀 Dynamic Tab Loop Engine
        for page_idx, document in enumerate(pages_data):
            
            # Setup Worksheet Tab
            if page_idx == 0:
                self.ws = self.wb.active
                self.ws.title = "Page 1"
            else:
                self.ws = self.wb.create_sheet(title=f"Page {page_idx + 1}")
                
            self.current_row = 1 # Reset Row count for the new page
            
            # --- RENDER THE PAGE ---
            max_cols = self.get_max_columns(document)
            
            # 🚀 DEFENSIVE MAIN TITLE
            main_title = document.get("main_title", {})
            if isinstance(main_title, str):
                main_title = {"text": main_title, "is_bold": True, "font_size": 14}
                
            self.write_merged_text(
                main_title.get("text", ""), max_cols, 
                main_title.get("is_bold", True), main_title.get("font_size", 14), 
                self.center_align
            )

            # 🚀 DEFENSIVE SUBTITLES
            subtitles = document.get("subtitles", [])
            if isinstance(subtitles, str):
                subtitles = [{"text": subtitles, "is_bold": True, "font_size": 12}]
            elif isinstance(subtitles, dict):
                subtitles = [subtitles]

            for subtitle in subtitles:
                if isinstance(subtitle, str):
                    subtitle = {"text": subtitle, "is_bold": True, "font_size": 12}
                    
                self.write_merged_text(
                    subtitle.get("text", ""), max_cols, 
                    subtitle.get("is_bold", True), subtitle.get("font_size", 12), 
                    self.center_align
                )
                
            self.current_row += 1 

            for table in document.get("tables", []):
                table_title = table.get("table_title", "")
                if table_title:
                    self.write_merged_text(table_title, max_cols, True, 12, self.left_align)

                headers = table.get("headers", [])
                for col_idx, header in enumerate(headers, start=1):
                    header_text = header.get("column_name", "")
                    cell = self.ws.cell(row=self.current_row, column=col_idx, value=self._process_text(header_text))
                    
                    cell.font = self._get_font(size=11, is_bold=header.get("is_bold", True))
                    cell.alignment = self.center_align
                    cell.border = self.thin_border
                    cell.fill = self.header_fill
                self.current_row += 1

                for row_data in table.get("rows", []):
                    max_lines_in_row = 1
                    for col_idx, value in enumerate(row_data, start=1):
                        cell = self.ws.cell(row=self.current_row, column=col_idx, value=self._process_text(str(value)))
                        
                        cell.font = self._get_font(size=11, is_bold=False)
                        cell.alignment = self.center_align
                        cell.border = self.thin_border
                        
                        lines = str(value).count('\n') + (len(str(value)) // 30) + 1
                        if lines > max_lines_in_row:
                            max_lines_in_row = lines
                            
                    self.ws.row_dimensions[self.current_row].height = max_lines_in_row * 16
                    self.current_row += 1
                    
                self.current_row += 1 

            footer = document.get("footer", {})
            if isinstance(footer, list):
                footer_text = "\n".join([str(i) for i in footer])
                footer = {"text": footer_text, "is_bold": False, "font_size": 11}
            elif isinstance(footer, str):
                footer = {"text": footer, "is_bold": False, "font_size": 11}
                
            self.write_merged_text(
                footer.get("text", ""), max_cols, 
                footer.get("is_bold", False), footer.get("font_size", 11), 
                self.left_align
            )

            # Autofit must happen per page BEFORE moving to the next tab!
            self._autofit_columns()

        self.wb.save(self.output_path)
        log.info(f"✅ Success! Smart Multi-Page Report saved to: {self.output_path}")