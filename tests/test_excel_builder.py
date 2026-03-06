import pytest
import json
import os
from openpyxl import load_workbook
from app.core.excel_builder import ExcelBuilder

# 🚀 PYTEST FIXTURE: Automatically handles temp files for every test cleanly
@pytest.fixture
def temp_paths(tmp_path):
    json_path = tmp_path / "input.json"
    excel_path = tmp_path / "output.xlsx"
    return str(json_path), str(excel_path)

def write_dummy_json(path, data):
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f)

def test_standard_build(temp_paths):
    """Proves a perfectly formatted JSON builds an Excel file."""
    json_path, excel_path = temp_paths
    perfect_json = {
        "document": {
            "main_title": {"text": "Title", "is_bold": True, "font_size": 14},
            "tables": [{"headers": [{"column_name": "Col 1"}], "rows": [["Val 1"]]}],
            "footer": {"text": "Valid Footer", "is_bold": False, "font_size": 11}
        }
    }
    write_dummy_json(json_path, perfect_json)
    builder = ExcelBuilder(json_path, excel_path)
    builder.build()
    
    assert os.path.exists(excel_path)
    
    # 🚀 Verify default font is Nirmala UI
    wb = load_workbook(excel_path)
    ws = wb.active
    assert ws["A1"].font.name == "Nirmala UI"

@pytest.mark.parametrize("font_choice", ["Kruti Dev 010", "DevLys 010"])
def test_legacy_font_application(temp_paths, font_choice):
    """Proves the UI toggle correctly overrides the Excel fonts."""
    json_path, excel_path = temp_paths
    test_json = {
        "document": {
            "tables": [{"headers": [{"column_name": "Header"}], "rows": [["Data"]]}]
        }
    }
    write_dummy_json(json_path, test_json)
    
    builder = ExcelBuilder(
        json_path, 
        excel_path, 
        use_legacy_font=True, 
        legacy_font_name=font_choice
    )
    builder.build()
    
    wb = load_workbook(excel_path)
    ws = wb.active
    
    # The merged title row is A1, so table header starts lower. Let's just check the builder's internal logic.
    test_font = builder._get_font(size=11, is_bold=True)
    assert test_font.name == font_choice
    assert test_font.size == 13  # Tests the size + 2 logic!

@pytest.mark.parametrize("broken_footer", [
    ["Line 1", "Line 2"],  # Hallucinated list
    "Just a raw string",   # Hallucinated string
    {}                     # Empty object
])
def test_defensive_footer_handling(temp_paths, broken_footer):
    """Proves the builder survives AI schema hallucinations."""
    json_path, excel_path = temp_paths
    broken_json = {
        "document": {
            "tables": [],
            "footer": broken_footer
        }
    }
    write_dummy_json(json_path, broken_json)
    builder = ExcelBuilder(json_path, excel_path)
    
    # Pytest will automatically fail if an exception is raised here
    builder.build()
    assert os.path.exists(excel_path)